"""
Django REST API for Excel file processing with Naver Shopping API price search
Supports card game price search and update functionality
"""

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status, serializers
import pandas as pd
import numpy as np
from django.http import HttpResponse, JsonResponse
from io import BytesIO
import os
import urllib.request
import urllib.parse
import json
import time
import re
import openpyxl
from openpyxl.styles import PatternFill
import logging
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import tempfile
from openpyxl.utils import get_column_letter

# API Configuration
NAVER_CLIENT_ID = "S_iul25XJKSybg_fiSAc"
NAVER_CLIENT_SECRET = "_73PsEM4om"
PLUS_PRICE = 0  # Additional amount to add to lowest price
API_DELAY = 0.3  # Delay between API requests (seconds)

# Excel Processing Configuration
PRODUCT_NAME_COLUMN = 3  # D column (0-indexed)
PRICE_COLUMN = 5  # F column (0-indexed)
DATA_START_ROW = 6  # Row where actual data starts

# Color definitions for price differences
COLOR_FILLS = {
    'none': PatternFill(fill_type=None),
    'green': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    'blue': PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),
    'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    'red': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
}

COLOR_LEGEND = [
    ("초록색", "1000원 이하", COLOR_FILLS['green']),
    ("파랑색", "2000원 이하", COLOR_FILLS['blue']),
    ("노랑색", "3000원 이하", COLOR_FILLS['yellow']),
    ("빨강색", "3000원 초과", COLOR_FILLS['red'])
]


class ExcelDataSerializer(serializers.Serializer):
    """Custom serializer for Excel data with proper null handling"""
    
    def to_representation(self, instance):
        data = {}
        for key, value in instance.items():
            if pd.isna(value) or (isinstance(value, (int, float)) and np.isinf(value)):
                data[key] = None
            elif isinstance(value, (np.int_, np.intc, np.intp, np.int8, np.int16, np.int32, np.int64)):
                data[key] = int(value)
            elif isinstance(value, (np.float16, np.float32, np.float64)):
                data[key] = None if (np.isnan(value) or np.isinf(value)) else float(value)
            else:
                data[key] = value
        return data


class CardGamePatternExtractor:
    """Card game pattern extraction and search keyword generation"""
    
    @staticmethod
    def extract_onepiece_info(product_name):
        """Extract One Piece card search information"""
        # 망가(슈퍼 패러렐) 패턴 체크 - 최우선 처리
        if "망가" in product_name:
            card_match = re.search(r'(OP|EB|ST)\d{2}-\d{3}', product_name)
            if card_match:
                card_number = card_match.group()
                return f"망가 {card_number}"
            else:
                return None
        
        # SP- 패턴 체크 (모두 스페셜 카드로 처리)
        sp_pattern = re.search(r'\bSP-(SP|SEC|R|SR|C|L|U|UC)\b', product_name)
        if sp_pattern:
            # 카드 번호 추출
            card_match = re.search(r'(OP|EB|ST)\d{2}-\d{3}', product_name)
            if card_match:
                card_number = card_match.group()
                return f"SP {card_number}"
            else:
                return None
        
        # P- 레어도 패턴 체크 (패러렐)
        has_p_rarity = bool(re.search(r'\bP-(SEC|R|SR|C|L|U)\b', product_name))
        
        # 카드 번호 패턴 찾기
        card_patterns = [
            (r'(OP|EB|ST)\d{2}-\d{3}', 'standard'),  # 일반 카드
            (r'P-\d{3}', 'promo')  # 프로모 카드
        ]
        
        for pattern, card_type in card_patterns:
            match = re.search(pattern, product_name)
            if match:
                card_number = match.group()
                
                if card_type == 'promo':
                    return f"원피스 {card_number}"
                elif has_p_rarity:
                    return f"패러렐 {card_number}"
                elif card_number.startswith('ST'):
                    return f"원피스 {card_number}"
                else:
                    return card_number
        
        # "원피스"로 시작하는 경우 추가 검색
        if product_name.startswith("원피스"):
            other_patterns = [
                (r'OP\d{2}-\d{3}', 'normal'),
                (r'(ST|EB|PR)\d{2}-\d{3}', 'special'),
                (r'P-\d{3}', 'promo')
            ]
            
            for pattern, ptype in other_patterns:
                match = re.search(pattern, product_name)
                if match:
                    card_number = match.group()
                    
                    if ptype == 'promo' or card_number.startswith('ST'):
                        result = f"원피스 {card_number}"
                    elif has_p_rarity:
                        result = f"패러렐 {card_number}"
                    else:
                        result = card_number
                    
                    return result
            
            # 등급 패턴 검색
            grade_match = re.search(r'(SR|R|C|L|SEC)\s+(OP|ST|EB|PR)\d{2}-\d{3}', product_name)
            if grade_match:
                card_number = grade_match.group(2)
                
                if has_p_rarity:
                    result = f"패러렐 {card_number}"
                elif card_number.startswith('ST'):
                    result = f"원피스 {card_number}"
                else:
                    result = card_number
                
                return result
        
        return None
    
    @staticmethod
    def extract_digimon_info(product_name):
        """Extract Digimon card search information"""
        # 새로운 형식: "디지몬카드"로 시작하는지 확인
        if not product_name.startswith("디지몬카드"):
            return None
        
        # 희소/패러렐 여부 확인
        has_rare = "희소" in product_name
        has_parallel = "패러렐" in product_name
        
        # 일반 카드 패턴
        card_match = re.search(r'(EX|BT|ST|RB|LM)\d{1,2}-\d{2,3}', product_name)
        if card_match:
            card_number = card_match.group()
            
            # 결과 결정
            is_st_card = card_number.startswith('ST')
            prefix = ""
            
            if has_rare:
                prefix = "희소 "
            elif has_parallel:
                prefix = "패러렐 "
            
            if is_st_card:
                result = f"{prefix}디지몬 {card_number}"
            else:
                result = f"{prefix}{card_number}" if prefix else card_number
            
            return result.strip()
        
        # 프로모 카드 패턴
        promo_match = re.search(r'P-\d{3}', product_name)
        if promo_match:
            card_number = promo_match.group()
            prefix = "희소 " if has_rare else ("패러렐 " if has_parallel else "")
            result = f"{prefix}디지몬 {card_number}"
            return result.strip()
        
        return None
    
    @staticmethod
    def extract_pokemon_info(product_name):
        """Extract Pokemon card search information"""
        if not product_name.startswith("포켓몬"):
            return None, None, None
        
        # 프로모 카드 확인
        promo_match = re.search(r'P-\d{3}', product_name)
        if promo_match:
            return f"포켓몬 {promo_match.group()}", None, None
        
        # 띄어쓰기로 구분 (마지막 단어=확장팩 제외)
        words = product_name.split()
        search_text = " ".join(words[:-1]) if len(words) > 1 else product_name
        last_word = words[-1] if len(words) > 1 else ""
        
        # 레어도 추출 - SSR 추가!
        rarity_pattern = r'\b(UR|SSR|SR|RR|RRR|CHR|CSR|BWR|AR|SAR|HR|R|U|C|몬스터볼|마스터볼|이로치)\b'
        rarity_match = re.search(rarity_pattern, search_text)
        rarity = rarity_match.group(1) if rarity_match else None
        
        # 포켓몬 이름 추출 (레어도 제거)
        temp_name = search_text
        if rarity:
            rarity_index = temp_name.find(rarity)
            if rarity_index != -1:
                temp_name = temp_name[:rarity_index].strip()
        
        # 특수 패턴 확인
        patterns = {
            'vmax': r'\b[가-힣A-Za-z\s]+(?:VMAX|Vmax|vmax)\b',
            'vstar': r'\b[가-힣A-Za-z\s]+(?:VStar|vstar|VSTAR)\b',
            'ex': r'\b[가-힣A-Za-z\s]+ex\b',
            'v': r'\b[가-힣A-Za-z\s]+V\b(?!\s*(?:MAX|max|Star|star))'
        }
        
        detected_patterns = {name: bool(re.search(pattern, temp_name, re.IGNORECASE)) 
                            for name, pattern in patterns.items()}
        
        # 포켓몬 이름 추출
        pokemon_name = None
        extraction_rules = [
            ('vmax', r'포켓몬카드\s+(.+?)\s*(?:VMAX|Vmax|vmax)'),
            ('vstar', r'포켓몬카드\s+(.+?)\s*(?:VStar|vstar|VSTAR)'),
            ('ex', r'포켓몬카드\s+(.+?ex)'),
            ('v', r'포켓몬카드\s+(.+?)\s*V\b(?!\s*(?:MAX|max|Star|star))'),
            (None, r'포켓몬카드\s+(.+)')
        ]
        
        for pattern_name, regex in extraction_rules:
            if pattern_name is None or detected_patterns.get(pattern_name, False):
                name_match = re.search(regex, temp_name, re.IGNORECASE)
                if name_match:
                    pokemon_name = name_match.group(1).strip()
                    break
        
        return product_name, rarity, pokemon_name
    
    @staticmethod
    def extract_search_info(product_name):
        """Extract search information from product name (unified function)"""
        # Try Digimon first (most specific pattern)
        digimon_result = CardGamePatternExtractor.extract_digimon_info(product_name)
        if digimon_result:
            return digimon_result, "디지몬", None
        
        # Try One Piece
        onepiece_result = CardGamePatternExtractor.extract_onepiece_info(product_name)
        if onepiece_result:
            return onepiece_result, "원피스", None
        
        # Try Pokemon
        pokemon_search, pokemon_rarity, pokemon_name = CardGamePatternExtractor.extract_pokemon_info(product_name)
        if pokemon_search:
            return pokemon_search, "포켓몬", (pokemon_rarity, pokemon_name)
        
        return None, None, None


class NaverShoppingAPI:
    """Naver Shopping API client"""
    
    @staticmethod
    def search(search_name):
        """Search Naver Shopping API"""
        try:
            enc_text = urllib.parse.quote(search_name)
            url = f"https://openapi.naver.com/v1/search/shop?query={enc_text}&sort=sim&exclude=used:rental:cbshop&display=20"
            
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
            request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
            
            response = urllib.request.urlopen(request)
            if response.getcode() == 200:
                result = json.loads(response.read())
                return result.get('items', [])
            else:
                logging.error("API request failed")
                return []
        except Exception as e:
            logging.error(f"API exception: {e}")
            return []


class ItemFilter:
    """Filter API search results based on card game rules"""
    
    @staticmethod
    def check_item_filters(title, mall_name, card_type, card_number,
                          is_parallel, is_rare, is_special_day, is_special,
                          is_super_parallel, price,
                          required_rarity, required_pokemon_name):
        """아이템 필터링 체크 (통과 여부와 로그 메시지 반환)"""
        
        # 제외 판매처
        if mall_name in ["화성스토어-TCG-", "네이버", "쿠팡"]:
            return False, f"❌ 제외: {mall_name}"
        
        # 일본판 제외
        if any(keyword in title for keyword in ['일본', '일본판', 'JP', 'JPN', '일판']):
            return False, "❌ 제외: 일본판"
        
        # 원피스/디지몬카드 번호 매칭
        if card_type in ["원피스", "디지몬"] and card_number:
            if card_number not in title:
                return False, f"❌ 제외: 카드번호 '{card_number}' 불일치"
        
        # 원피스 슈퍼 패러렐(망가) 키워드 확인
        if card_type == "원피스" and is_super_parallel:
            super_parallel_keywords = ['슈퍼 패러렐', '슈퍼패러렐', '슈퍼파라렐', '슈퍼 파라렐']
            manga_keywords = ['망가', 'MANGA', 'manga']
            
            # 슈퍼 패러렐 키워드 확인
            has_super_parallel = any(kw in title for kw in super_parallel_keywords)
            # 망가 키워드 확인
            has_manga = any(kw in title for kw in manga_keywords)
            
            # 두 키워드 중 하나라도 포함되어야 함
            if not (has_super_parallel or has_manga):
                return False, "❌ 제외: 슈퍼 패러렐 또는 망가 키워드 없음"
            
            # 가격 체크: 200,000원 미만 제외
            if price < 200000:
                return False, f"❌ 제외: 가격 {int(price)}원 (20만원 미만)"
            
            matched_keywords = []
            if has_super_parallel:
                matched_keywords.append("슈퍼 패러렐")
            if has_manga:
                matched_keywords.append("망가")
            
            logging.info(f"✅ 슈퍼 패러렐(망가) 키워드 매칭 성공 ({', '.join(matched_keywords)}) (가격: {int(price)}원)")
        
        # 원피스 스페셜 키워드 확인
        elif card_type == "원피스" and is_special:
            special_keywords = ['스페셜', 'SP']
            matched_keyword = next((kw for kw in special_keywords if kw in title), None)
            if not matched_keyword:
                return False, "❌ 제외: 스페셜 키워드 없음"
        
        # 원피스 패러렐 키워드 확인
        elif card_type == "원피스" and is_parallel:
            parallel_keywords = ['패러렐', '다른', '패레', 'P시크릿레어', '페러럴', '패러럴', '페러렐', '페레']
            matched_keyword = next((kw for kw in parallel_keywords if kw in title), None)
            if not matched_keyword:
                return False, "❌ 제외: 패러렐 키워드 없음"
        
        # 디지몬 희소/패러렐 키워드 확인
        elif card_type == "디지몬":
            if is_rare and "희소" not in title:
                return False, "❌ 제외: 희소 키워드 없음"
            
            if is_parallel and "패러렐" not in title:
                return False, "❌ 제외: 패러렐 키워드 없음"
        
        # 포켓몬카드 조건 확인
        elif card_type == "포켓몬":
            if is_special_day and "특일" not in title:
                return False, "❌ 제외: 특일 키워드 없음"
            
            # 포켓몬 이름 매칭
            if required_pokemon_name:
                clean_title = re.sub(r'<[^>]+>', '', title)
                
                # 띄어쓰기 제거 매칭
                required_name_no_space = re.sub(r'\s+', '', required_pokemon_name)
                title_no_space = re.sub(r'\s+', '', clean_title)
                
                if required_name_no_space.lower() in title_no_space.lower():
                    pass  # 매칭 성공
                else:
                    # 개별 단어 매칭
                    required_words = [word for word in required_pokemon_name.split() 
                                    if word.lower() not in ['ex', 'v', 'vmax', 'vstar']]
                    
                    word_matches = sum(1 for word in required_words if word.lower() in clean_title.lower())
                    
                    if word_matches != len(required_words) or len(required_words) == 0:
                        return False, f"❌ 개별 단어 매칭 실패 ({word_matches}/{len(required_words)})"
            
            # 레어도 매칭
            if required_rarity:
                clean_title = re.sub(r'<[^>]+>', '', title)
                
                if required_rarity not in clean_title:
                    return False, f"❌ 제외: 레어도 '{required_rarity}' 미포함"
        
        return True, "✅ 통과: 필터링 조건 만족"
    
    @staticmethod
    def filter_api_results(items, search_name, card_type, pokemon_info=None):
        """API 검색 결과 필터링"""
        min_price = None
        valid_items_count = 0
        filter_match_info = "없음"
        
        # 검색 조건 설정
        is_super_parallel = "망가" in search_name  # 망가 키워드로 체크
        is_parallel = "패러렐" in search_name and not is_super_parallel
        is_rare = "희소" in search_name
        is_special_day = "특일" in search_name
        is_special = "SP" in search_name and not is_super_parallel
        
        # 카드별 기본 필터 정보
        if card_type == "원피스":
            if is_super_parallel:
                filter_match_info = "슈퍼패러렐(망가)검색"
            elif is_special:
                filter_match_info = "스페셜검색"
            elif is_parallel:
                filter_match_info = "패러렐검색"
            else:
                filter_match_info = "일반검색"
        elif card_type == "디지몬":
            filter_match_info = "희소검색" if is_rare else ("패러렐검색" if is_parallel else "일반검색")
        elif card_type == "포켓몬":
            filter_match_info = "필터없음"
        
        # 카드 번호 추출
        card_number = None
        if card_type in ["원피스", "디지몬"]:
            pattern = r'(OP|ST|EB|PR)\d{2}-\d{3}' if card_type == "원피스" else r'(EX|BT|ST|RB|LM)\d{1,2}-\d{3}'
            card_match = re.search(pattern, search_name)
            card_number = card_match.group() if card_match else None
        
        # 포켓몬카드 정보
        required_rarity, required_pokemon_name = pokemon_info or (None, None)
        
        # 아이템 필터링
        for item in items:
            title = item['title']
            price = float(item['lprice'])
            mall_name = item.get('mallName', '')
            
            # 필터 체크
            passed, log_msg = ItemFilter.check_item_filters(
                title, mall_name, card_type, card_number,
                is_parallel, is_rare, is_special_day, is_special,
                is_super_parallel, price,
                required_rarity, required_pokemon_name
            )
            
            if not passed:
                continue
            
            # 통과한 상품
            valid_items_count += 1
            
            # 최저가 업데이트
            if min_price is None or price < min_price:
                min_price = price
                
                # 포켓몬카드 필터 정보 업데이트
                if card_type == "포켓몬":
                    if required_pokemon_name and required_rarity:
                        filter_match_info = "포켓몬명+레어도"
                    elif required_pokemon_name:
                        filter_match_info = "포켓몬명만"
                    elif required_rarity:
                        filter_match_info = "레어도만"
                    else:
                        filter_match_info = "필터없음"
        
        return min_price, valid_items_count, filter_match_info


class PriceProcessor:
    """Process price updates for card games"""
    
    @staticmethod
    def process_price_update(product_name, original_price):
        """가격 업데이트 처리"""
        search_name, card_type, pokemon_info = CardGamePatternExtractor.extract_search_info(product_name)
        
        if not search_name:
            logging.info(f"{product_name} : {int(original_price)} (검색 패턴 없음)")
            return original_price, 0, "미확인", "패턴없음", "패턴없음", 0
        
        # API 검색
        items = NaverShoppingAPI.search(search_name)
        min_price, valid_items_count, filter_match_info = ItemFilter.filter_api_results(
            items, search_name, card_type, pokemon_info
        )
        
        # 가격 계산
        new_price = (min_price + PLUS_PRICE) if min_price is not None else original_price
        price_diff = int(new_price - original_price)
        
        # 상세 로깅
        if abs(price_diff) > 0:
            if card_type == "포켓몬" and pokemon_info:
                rarity, pokemon_name = pokemon_info
                info_text = f" (포켓몬: {pokemon_name or '없음'}"
                if rarity:
                    info_text += f", 레어도: {rarity}"
                info_text += f", 필터: {filter_match_info})"
                logging.info(f"{product_name} : {int(original_price)} → {int(new_price)} ({price_diff:+}원) [{card_type}카드 전체 검색{info_text}]")
            else:
                logging.info(f"{product_name} : {int(original_price)} → {int(new_price)} ({price_diff:+}원) [{card_type}카드 검색어: {search_name}]")
        else:
            if card_type == "포켓몬" and pokemon_info:
                rarity, pokemon_name = pokemon_info
                info_text = f" (포켓몬: {pokemon_name or '없음'}"
                if rarity:
                    info_text += f", 레어도: {rarity}"
                info_text += f", 필터: {filter_match_info})"
                logging.info(f"{product_name} : {int(original_price)} (변경없음) [{card_type}카드 전체 검색{info_text}]")
            else:
                logging.info(f"{product_name} : {int(original_price)} (변경없음) [{card_type}카드 검색어: {search_name}]")
        
        logging.info("-" * 60)
        
        time.sleep(API_DELAY)  # API 요청 제한 방지
        return new_price, price_diff, card_type, filter_match_info, search_name, valid_items_count
    
    @staticmethod
    def get_fill_color(original_price, new_price):
        """가격 차이에 따른 색상 결정"""
        if abs(original_price - new_price) < 0.01:
            return COLOR_FILLS['none']
        
        price_diff = abs(new_price - original_price)
        
        if price_diff <= 1000:
            return COLOR_FILLS['green']
        elif price_diff <= 2000:
            return COLOR_FILLS['blue']
        elif price_diff <= 3000:
            return COLOR_FILLS['yellow']
        else:
            return COLOR_FILLS['red']


# ==================== API Endpoints ====================

# 로깅 설정 - 콘솔에도 출력되도록 설정
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 콘솔 핸들러 추가
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(message)s')  # 심플한 포맷
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_excel(request):
    """
    Upload Excel file and extract data
    
    Expected file format:
    - D column: Product names (상품명)
    - F column: Prices (가격)
    - Data starts from row 6
    """
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Invalid file format. Please upload .xlsx or .xls file'}, 
                       status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, header=None)
        
        # Extract relevant columns (D=3, F=5, 0-indexed)
        # Get all rows starting from index 5 (6th row in Excel, accounting for 0-indexing)
        data_rows = []
        for idx in range(DATA_START_ROW - 1, len(df)):  # -1 for 0-indexing
            product_name = df.iloc[idx, PRODUCT_NAME_COLUMN]
            price = df.iloc[idx, PRICE_COLUMN]
            
            # Skip if both values are NaN
            if pd.isna(product_name) and pd.isna(price):
                continue
            
            data_rows.append({
                'excelRow': idx + 1,  # Convert to 1-indexed Excel row number
                'productName': None if pd.isna(product_name) else str(product_name),
                'price': None if pd.isna(price) else float(price)
            })
        
        # Serialize data properly handling NaN/None values
        serializer = ExcelDataSerializer(data_rows, many=True)
        
        return Response({
            'message': 'File uploaded successfully',
            'data': serializer.data,
            'totalRows': len(data_rows)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f'Failed to process file: {str(e)}'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def search_prices(request):
    """
    Search prices for card game products using Naver Shopping API
    
    Request body:
    {
        "items": [
            {"productName": "string", "currentPrice": float},
            ...
        ]
    }
    
    Response:
    {
        "results": [
            {
                "productName": "string",
                "currentPrice": float,
                "newPrice": float,
                "priceDiff": int,
                "cardType": "string",
                "filterInfo": "string",
                "searchKeyword": "string",
                "validItemsCount": int
            },
            ...
        ]
    }
    """
    try:
        items = request.data.get('items', [])
        
        if not items:
            return Response({'error': 'No items provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 시작 로그
        logging.info("=" * 80)
        logging.info("🚀 카드 최저가 검색 시작")
        logging.info("=" * 80)
        logging.info(f"처리할 상품 수: {len(items)}개")
        logging.info(f"현재 최저가에서 {PLUS_PRICE}원 추가됩니다.\n")
        
        results = []
        
        for idx, item in enumerate(items, 1):
            product_name = item.get('productName')
            current_price = item.get('currentPrice', 0)
            
            if not product_name:
                continue
            
            logging.info(f"[{idx}/{len(items)}] 처리 중...")
            
            try:
                new_price, price_diff, card_type, filter_info, search_keyword, valid_count = \
                    PriceProcessor.process_price_update(product_name, float(current_price))
                
                results.append({
                    'productName': product_name,
                    'currentPrice': current_price,
                    'newPrice': new_price,
                    'priceDiff': price_diff,
                    'cardType': card_type,
                    'filterInfo': filter_info,
                    'searchKeyword': search_keyword,
                    'validItemsCount': valid_count
                })
            except Exception as e:
                logging.error(f"상품 처리 중 오류 ({product_name}): {str(e)}")
                results.append({
                    'productName': product_name,
                    'currentPrice': current_price,
                    'newPrice': current_price,
                    'priceDiff': 0,
                    'cardType': '오류',
                    'filterInfo': '처리실패',
                    'searchKeyword': '처리실패',
                    'validItemsCount': 0,
                    'error': str(e)
                })
        
        # 완료 로그
        logging.info("\n" + "=" * 80)
        logging.info("✅ 카드 최저가 검색 완료")
        logging.info("=" * 80)
        changed_count = sum(1 for r in results if r['priceDiff'] != 0)
        logging.info(f"총 {len(results)}개 상품 처리 완료")
        logging.info(f"가격 변경: {changed_count}개")
        logging.info(f"변경 없음: {len(results) - changed_count}개\n")
        
        return Response({
            'results': results,
            'totalProcessed': len(results)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logging.error(f"가격 검색 중 오류 발생: {str(e)}")
        return Response({'error': f'Failed to search prices: {str(e)}'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@csrf_exempt
@require_http_methods(["POST"])
def download_excel(request):
    """
    Download modified Excel file with updated prices and stock
    추가 정보: A~F열에 변동액, 기존가격, 카드타입, 필터적용, 검색개수, 검색어 추가
    가격 변동에 따라 색상 적용
    
    Request body (JSON):
    {
        "modifications": [
            {
                "excelRow": int,
                "price": float,
                "stock": int,
                "productName": "string" (optional, for logging)
            },
            ...
        ]
    }
    
    File upload (multipart/form-data):
    - "excel_file": Excel file (.xlsx)
    """
    temp_file_path = None
    output_temp_path = None
    
    try:
        # 1. 로깅 설정
        logger.info("=" * 50)
        logger.info("Excel 파일 처리 시작")
        logger.info("=" * 50)
        
        # 2. 요청 데이터 파싱
        if 'excel_file' not in request.FILES:
            return JsonResponse({'error': '파일이 제공되지 않았습니다'}, status=400)
        
        excel_file = request.FILES['excel_file']
        original_filename = excel_file.name
        logger.info(f"업로드된 파일: {original_filename}")
        logger.info(f"파일 크기: {excel_file.size} bytes")
        
        # modifications 데이터 파싱
        try:
            modifications_json = request.POST.get('modifications')
            if not modifications_json:
                return JsonResponse({'error': 'modifications 데이터가 없습니다'}, status=400)
            
            modifications = json.loads(modifications_json)
            logger.info(f"수정 항목 개수: {len(modifications)}")
        except json.JSONDecodeError as e:
            logger.error(f"JSON 파싱 오류: {str(e)}")
            return JsonResponse({'error': 'modifications JSON 파싱 실패'}, status=400)
        
        # 3. 임시 파일 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file_path = temp_file.name
            for chunk in excel_file.chunks():
                temp_file.write(chunk)
        
        logger.info(f"임시 파일 생성: {temp_file_path}")
        
        # 4. 엑셀 파일 로드
        try:
            workbook = openpyxl.load_workbook(temp_file_path)
            worksheet = workbook.worksheets[0]
            logger.info(f"워크북 로드 성공")
        except Exception as e:
            logger.error(f"워크북 로드 실패: {str(e)}")
            raise e
        
        # 5. 기존 워크시트를 읽어서 새 워크북에 A~F열 추가하여 재구성
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        
        logger.info("=" * 30)
        logger.info("새 워크시트 생성 - A~F열 추가")
        logger.info("=" * 30)
        
        # 6. modifications를 딕셔너리로 변환 (빠른 조회를 위해)
        mod_dict = {int(mod['excelRow']): mod for mod in modifications}
        
        # 7. 모든 행 처리
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            new_row = []
            price_info = None
            
            # 첫 번째 행 (헤더)
            if row_idx == 1:
                new_row.extend(["변동액", "기존가격", "카드타입", "필터적용", "검색개수", "검색어"])
                # 기존 데이터 추가
                for cell in row:
                    new_row.append(cell.value)
            else:
                # 데이터 행 - 기본값 설정
                new_row.extend([0, 0, "", "", 0, ""])
                
                # 수정 정보가 있는 경우
                if row_idx in mod_dict:
                    mod = mod_dict[row_idx]
                    product_name = mod.get('productName', '')
                    
                    # D열(상품명)에서 원본 가격 가져오기
                    original_price_cell = worksheet.cell(row=row_idx, column=6)  # F열
                    original_price = float(original_price_cell.value) if original_price_cell.value else 0
                    new_price = float(mod.get('price', original_price))
                    price_diff = int(new_price - original_price)
                    
                    # 카드 정보 추출 (최저가 검색 시 저장된 정보)
                    search_name, card_type, pokemon_info = CardGamePatternExtractor.extract_search_info(product_name)
                    
                    # A~F열 정보 설정
                    new_row[0] = price_diff  # 변동액
                    new_row[1] = int(original_price)  # 기존가격
                    new_row[2] = card_type or "미확인"  # 카드타입
                    new_row[3] = mod.get('filterInfo', "")  # 필터적용 (프론트에서 전달)
                    new_row[4] = mod.get('validCount', 0)  # 검색개수 (프론트에서 전달)
                    new_row[5] = search_name or ""  # 검색어
                    
                    price_info = (original_price, new_price)
                    
                    logger.info(f"행 {row_idx}: {product_name} | {int(original_price)} → {int(new_price)} ({price_diff:+}원)")
                
                # 기존 데이터 복사
                for cell in row:
                    # F열(가격) 또는 H열(재고)이고 수정 정보가 있으면 새 값 사용
                    if row_idx in mod_dict:
                        mod = mod_dict[row_idx]
                        if cell.column == 6:  # F열 (가격)
                            new_row.append(float(mod.get('price', cell.value or 0)))
                        elif cell.column == 8:  # H열 (재고)
                            new_row.append(int(float(mod.get('stock', cell.value or 0))))
                        else:
                            new_row.append(cell.value)
                    else:
                        new_row.append(cell.value)
            
            # 새 워크시트에 행 추가
            new_worksheet.append(new_row)
            
            # 가격 셀에 색상 적용 (A~F 6개 컬럼 추가되어 F열이 12열로 이동)
            if price_info is not None and row_idx > 1:
                price_cell = new_worksheet.cell(row=row_idx, column=12)  # F열이 12열로 이동
                fill_color = PriceProcessor.get_fill_color(price_info[0], price_info[1])
                price_cell.fill = fill_color
        
        # 8. 색상 범례 추가 (첫 번째 열의 2~5행)
        color_legend = [
            ("초록색", "1000원 이하", COLOR_FILLS['green']),
            ("파랑색", "2000원 이하", COLOR_FILLS['blue']),
            ("노랑색", "3000원 이하", COLOR_FILLS['yellow']),
            ("빨강색", "3000원 초과", COLOR_FILLS['red'])
        ]
        
        for i, (color_name, range_text, fill_color) in enumerate(color_legend, 2):
            new_worksheet.cell(row=i, column=1, value=color_name).fill = fill_color
            new_worksheet.cell(row=i, column=2, value=range_text)
        
        logger.info("색상 범례 추가 완료")
        
        # 9. 파일 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as output_temp:
            output_temp_path = output_temp.name
        
        try:
            new_workbook.save(output_temp_path)
            logger.info("새 워크북 저장 완료")
        except Exception as e:
            logger.error(f"워크북 저장 실패: {str(e)}")
            raise e
        finally:
            new_workbook.close()
            workbook.close()
        
        # 10. 저장된 파일 검증
        output_size = os.path.getsize(output_temp_path)
        logger.info(f"저장된 파일 크기: {output_size} bytes")
        
        if output_size == 0:
            raise Exception("저장된 파일 크기가 0입니다")
        
        # 11. HTTP 응답 생성
        with open(output_temp_path, 'rb') as f:
            file_content = f.read()
        
        base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
        new_filename = f"{base_name}_수정.xlsx"
        
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{new_filename}"'
        response['Content-Length'] = len(file_content)
        
        logger.info("=" * 50)
        logger.info("Excel 파일 처리 완료")
        logger.info("=" * 50)
        logger.info(f"파일명: {new_filename}")
        logger.info(f"응답 크기: {len(file_content)} bytes")
        logger.info(f"\n추가된 정보:")
        logger.info(f"   A열: 변동액 (정수)")
        logger.info(f"   B열: 기존가격")
        logger.info(f"   C열: 카드 타입")
        logger.info(f"   D열: 필터 적용 여부")
        logger.info(f"        - 원피스: 일반검색/패러렐검색/스페셜검색/슈퍼패러렐(망가)검색")
        logger.info(f"        - 디지몬: 일반검색/희소검색/패러렐검색")
        logger.info(f"        - 포켓몬: 포켓몬명만/포켓몬명+레어도/레어도만/필터없음")
        logger.info(f"   E열: 검색된 상품 개수")
        logger.info(f"   F열: 검색어")
        
        return response
        
    except Exception as e:
        logger.error("=" * 50)
        logger.error("전체 처리 중 오류 발생")
        logger.error("=" * 50)
        logger.error(f"오류: {str(e)}")
        import traceback
        logger.error(f"스택 트레이스:\n{traceback.format_exc()}")
        return JsonResponse({'error': f'처리 중 오류가 발생했습니다: {str(e)}'}, status=500)
        
    finally:
        # 임시 파일 정리
        for file_path, desc in [(temp_file_path, "입력"), (output_temp_path, "출력")]:
            if file_path and os.path.exists(file_path):
                try:
                    os.unlink(file_path)
                    logger.info(f"{desc} 임시 파일 삭제: {file_path}")
                except Exception as e:
                    logger.warning(f"{desc} 임시 파일 삭제 실패: {e}")