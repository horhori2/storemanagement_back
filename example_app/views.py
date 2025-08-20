"""
Django REST API for Excel file processing with Naver Shopping API price search
Supports card game price search and update functionality
"""

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status, serializers
import pandas as pd
import numpy as np
from django.http import HttpResponse
from io import BytesIO
import os
import urllib.request
import urllib.parse
import json
import time
import re
import openpyxl
from openpyxl.styles import PatternFill

# API Configuration
NAVER_CLIENT_ID = "S_iul25XJKSybg_fiSAc"
NAVER_CLIENT_SECRET = "_73PsEM4om"
PLUS_PRICE = 0  # Additional amount to add to lowest price
API_DELAY = 0.3  # Delay between API requests (seconds)

# Excel Processing Configuration
PRODUCT_NAME_COLUMN = 3  # D column (0-indexed)
PRICE_COLUMN = 5  # F column (0-indexed)
DATA_START_ROW = 6  # Row where actual data starts

# Color definitions for price differences
COLOR_FILLS = {
    'none': PatternFill(fill_type=None),
    'green': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    'blue': PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),
    'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    'red': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
}

COLOR_LEGEND = [
    ("초록색", "1000원 이하", COLOR_FILLS['green']),
    ("파랑색", "2000원 이하", COLOR_FILLS['blue']),
    ("노랑색", "3000원 이하", COLOR_FILLS['yellow']),
    ("빨강색", "3000원 초과", COLOR_FILLS['red'])
]


class ExcelDataSerializer(serializers.Serializer):
    """Custom serializer for Excel data with proper null handling"""
    
    def to_representation(self, instance):
        data = {}
        for key, value in instance.items():
            if pd.isna(value) or (isinstance(value, (int, float)) and np.isinf(value)):
                data[key] = None
            elif isinstance(value, (np.int_, np.intc, np.intp, np.int8, np.int16, np.int32, np.int64)):
                data[key] = int(value)
            elif isinstance(value, (np.float16, np.float32, np.float64)):
                data[key] = None if (np.isnan(value) or np.isinf(value)) else float(value)
            else:
                data[key] = value
        return data


class CardGamePatternExtractor:
    """Card game pattern extraction and search keyword generation"""
    
    @staticmethod
    def extract_onepiece_info(product_name):
        """Extract One Piece card search information"""
        # Exclude SP- rarity patterns
        if re.search(r'\bSP-(SEC|R|SR|C|L|U|SP)\b', product_name):
            return "EXCLUDE"
        
        # Check for P- rarity (parallel processing)
        has_p_rarity = bool(re.search(r'\bP-(SEC|R|SR|C|L|U)\b', product_name))
        
        # Card number patterns
        card_patterns = [
            r'(OP|EB|ST)\d{2}-\d{3}',  # Regular cards
            r'P-\d{3}'  # Promo cards
        ]
        
        for pattern in card_patterns:
            match = re.search(pattern, product_name)
            if match:
                card_number = match.group()
                if pattern == r'P-\d{3}':
                    return f"원피스 {card_number}"
                return f"패러렐 {card_number}" if has_p_rarity else card_number
        
        # Handle products starting with "원피스"
        if product_name.startswith("원피스"):
            return CardGamePatternExtractor._extract_onepiece_fallback(product_name, has_p_rarity)
        
        return None
    
    @staticmethod
    def _extract_onepiece_fallback(product_name, has_p_rarity):
        """Fallback extraction for One Piece cards starting with '원피스'"""
        other_patterns = [r'OP\d{2}-\d{3}', r'(ST|EB|PR)\d{2}-\d{3}', r'P-\d{3}']
        
        for pattern in other_patterns:
            match = re.search(pattern, product_name)
            if match:
                card_number = match.group()
                if pattern == r'P-\d{3}':
                    return f"원피스 {card_number}"
                return f"패러렐 {card_number}" if has_p_rarity else card_number
        
        # Grade and number pattern
        grade_match = re.search(r'(SR|R|C|L|SEC)\s+(OP|ST|EB|PR)\d{2}-\d{3}', product_name)
        if grade_match:
            card_number = grade_match.group(2)
            return f"패러렐 {card_number}" if has_p_rarity else card_number
        
        return None
    
    @staticmethod
    def extract_digimon_info(product_name):
        """Extract Digimon card search information"""
        parts = product_name.split("/")
        
        if not parts or not parts[-1].strip().startswith("디지몬"):
            return None
        
        has_rare = any("희소" in part for part in parts)
        has_parallel = any("패러렐" in part for part in parts)
        
        if len(parts) >= 2:
            code_part = parts[-2].strip()
            
            # Regular card pattern
            digimon_match = re.search(r'(EX|BT|ST|RB|LM)\d{1,2}-\d{3}', code_part)
            if digimon_match:
                card_number = digimon_match.group()
                if has_rare:
                    return f"희소 {card_number}"
                elif has_parallel:
                    return f"패러렐 {card_number}"
                else:
                    return card_number
            
            # Promo card pattern
            promo_match = re.search(r'P-\d{3}', code_part)
            if promo_match:
                card_number = promo_match.group()
                prefix = "희소 디지몬" if has_rare else ("패러렐 디지몬" if has_parallel else "디지몬")
                return f"{prefix} {card_number}"
        
        return None
    
    @staticmethod
    def extract_pokemon_info(product_name):
        """Extract Pokemon card search information"""
        if not product_name.startswith("포켓몬"):
            return None, None, None
        
        # Promo card check
        promo_match = re.search(r'P-\d{3}', product_name)
        if promo_match:
            return f"포켓몬 {promo_match.group()}", None, None
        
        # Extract rarity
        rarity_pattern = r'\b(UR|SR|RR|RRR|CHR|CSR|BWR|AR|SAR|R|U|C|몬스터볼|마스터볼)\b'
        rarity_match = re.search(rarity_pattern, product_name)
        rarity = rarity_match.group(1) if rarity_match else None
        
        # Extract Pokemon name
        name_match = re.search(r'포켓몬카드\s+([가-힣A-Za-z]+)', product_name)
        pokemon_name = name_match.group(1) if name_match else None
        
        return product_name, rarity, pokemon_name
    
    @staticmethod
    def extract_search_info(product_name):
        """Extract search information from product name (unified function)"""
        # Try Digimon first (most specific pattern)
        digimon_result = CardGamePatternExtractor.extract_digimon_info(product_name)
        if digimon_result:
            return digimon_result, "디지몬", None
        
        # Try One Piece
        onepiece_result = CardGamePatternExtractor.extract_onepiece_info(product_name)
        if onepiece_result == "EXCLUDE":
            return "EXCLUDE", "원피스", None
        elif onepiece_result:
            return onepiece_result, "원피스", None
        
        # Try Pokemon
        pokemon_search, pokemon_rarity, pokemon_name = CardGamePatternExtractor.extract_pokemon_info(product_name)
        if pokemon_search:
            return pokemon_search, "포켓몬", (pokemon_rarity, pokemon_name)
        
        return None, None, None


class NaverShoppingAPI:
    """Naver Shopping API client"""
    
    @staticmethod
    def search(search_name):
        """Search products using Naver Shopping API"""
        try:
            enc_text = urllib.parse.quote(search_name)
            url = f"https://openapi.naver.com/v1/search/shop?query={enc_text}&sort=sim&exclude=used:rental:cbshop&display=10"
            
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
            request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
            
            response = urllib.request.urlopen(request)
            if response.getcode() == 200:
                result = json.loads(response.read())
                return result.get('items', [])
            else:
                print("❌ API 요청 실패")
                return []
        except Exception as e:
            print(f"❌ 예외 발생: {e}")
            return []
    
    @staticmethod
    def filter_results(items, search_name, card_type, pokemon_info=None):
        """Filter API search results based on card type and conditions"""
        min_price = None
        is_parallel = "패러렐" in search_name
        is_rare = "희소" in search_name
        
        # Extract card number
        card_number = NaverShoppingAPI._extract_card_number(search_name, card_type, is_parallel, is_rare)
        required_rarity, required_pokemon_name = pokemon_info or (None, None)
        
        for item in items:
            title = item['title']
            price = float(item['lprice'])
            
            # Skip Japanese versions
            if any(keyword in title for keyword in ['일본', '일본판', 'JP', 'JPN']):
                continue
            
            # Card number matching for One Piece/Digimon
            if card_type in ["원피스", "디지몬"] and card_number and card_number not in title:
                continue
            
            # Special card keyword checks
            if not NaverShoppingAPI._check_special_conditions(title, card_type, is_parallel, is_rare):
                continue
            
            # Pokemon card conditions
            if card_type == "포켓몬" and not NaverShoppingAPI._check_pokemon_conditions(
                title, required_pokemon_name, required_rarity
            ):
                continue
            
            # Update minimum price
            if min_price is None or price < min_price:
                min_price = price
        
        return min_price
    
    @staticmethod
    def _extract_card_number(search_name, card_type, is_parallel, is_rare):
        """Extract card number from search name"""
        if card_type == "원피스":
            if is_parallel:
                card_match = re.search(r'(OP|ST|EB|PR)\d{2}-\d{3}', search_name)
                return card_match.group() if card_match else None
            elif re.match(r'(OP|ST|EB|PR)\d{2}-\d{3}', search_name):
                return search_name
        elif card_type == "디지몬":
            if is_parallel or is_rare:
                card_match = re.search(r'(EX|BT|ST|RB|LM)\d{1,2}-\d{3}', search_name)
                return card_match.group() if card_match else None
            elif re.match(r'(EX|BT|ST|RB|LM)\d{1,2}-\d{3}', search_name):
                return search_name
        return None
    
    @staticmethod
    def _check_special_conditions(title, card_type, is_parallel, is_rare):
        """Check special conditions for card types"""
        if card_type == "원피스" and is_parallel:
            return any(keyword in title for keyword in [
                '패러렐', '다른', '패레', 'P시크릿레어', '페러럴', '패러럴', '페러렐', '페레'
            ])
        elif card_type == "디지몬":
            if is_rare and "희소" not in title:
                return False
            if is_parallel and "패러렐" not in title:
                return False
        return True
    
    @staticmethod
    def _check_pokemon_conditions(title, required_pokemon_name, required_rarity):
        """Check Pokemon card specific conditions"""
        if required_pokemon_name and required_pokemon_name not in title:
            return False
        
        if required_rarity:
            title_rarity = re.search(r'\b(UR|SR|RR|RRR|CHR|CSR|BWR|AR|SAR|R|U|C|몬스터볼|마스터볼)\b', title)
            if not title_rarity or title_rarity.group(1) != required_rarity:
                return False
        return True


class PriceProcessor:
    """Price processing and updating logic"""
    
    @staticmethod
    def update_price(product_name, original_price):
        """Update price based on Naver API search"""
        search_name, card_type, pokemon_info = CardGamePatternExtractor.extract_search_info(product_name)
        
        if search_name == "EXCLUDE":
            print(f"{product_name} : {int(original_price)} (SP- 레어도 - 변경없음)")
            return original_price, "0원"
        
        if not search_name:
            print(f"{product_name} : {int(original_price)} (검색 패턴 없음)")
            return original_price, "0원"
        
        # API search
        items = NaverShoppingAPI.search(search_name)
        min_price = NaverShoppingAPI.filter_results(items, search_name, card_type, pokemon_info)
        
        # Calculate new price
        new_price = (min_price + PLUS_PRICE) if min_price is not None else original_price
        price_diff = new_price - original_price
        change_text = f"{price_diff:+.0f}원" if abs(price_diff) > 0.01 else "0원"
        
        # Log the change
        PriceProcessor._log_price_change(product_name, original_price, new_price, price_diff, card_type, search_name)
        
        time.sleep(API_DELAY)  # Rate limiting
        return new_price, change_text
    
    @staticmethod
    def _log_price_change(product_name, original_price, new_price, price_diff, card_type, search_name):
        """Log price change information"""
        if abs(price_diff) > 0.01:
            print(f"{product_name} : {int(original_price)} → {int(new_price)} "
                  f"({price_diff:+.0f}원) [{card_type}카드 검색어: {search_name}]")
        else:
            print(f"{product_name} : {int(original_price)} (변경없음) [{card_type}카드 검색어: {search_name}]")
    
    @staticmethod
    def get_color_fill(original_price, new_price):
        """Determine color based on price difference"""
        if abs(original_price - new_price) < 0.01:
            return COLOR_FILLS['none']
        
        price_diff = abs(new_price - original_price)
        
        if price_diff <= 1000:
            return COLOR_FILLS['green']
        elif price_diff <= 2000:
            return COLOR_FILLS['blue']
        elif price_diff <= 3000:
            return COLOR_FILLS['yellow']
        else:
            return COLOR_FILLS['red']


class ExcelProcessor:
    """Excel file processing and manipulation"""
    
    @staticmethod
    def add_color_legend(worksheet):
        """Add color legend to A2~B5"""
        for i, (color_name, range_text, fill_color) in enumerate(COLOR_LEGEND, 2):
            worksheet.cell(row=i, column=1, value=color_name).fill = fill_color
            worksheet.cell(row=i, column=2, value=range_text)
    
    @staticmethod
    def process_workbook(excel_file):
        """Process Excel workbook with price updates and formatting"""
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        price_changes = []
        
        # Process each row
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            new_row, price_info, change_info = ExcelProcessor._process_row(row, row_idx)
            new_ws.append(new_row)
            
            if price_info:
                # Apply color to price cell (column H after adding A,B columns)
                price_cell = new_ws.cell(row=new_ws.max_row, column=8)
                price_cell.fill = PriceProcessor.get_color_fill(price_info[0], price_info[1])
            
            if change_info:
                price_changes.append(change_info)
                print("-" * 60)
        
        # Add color legend
        ExcelProcessor.add_color_legend(new_ws)
        
        return new_wb, price_changes
    
    @staticmethod
    def _process_row(row, row_idx):
        """Process individual row"""
        new_row = []
        price_info = None
        change_info = None
        
        # Add headers for A and B columns
        if row_idx == 1:
            new_row.extend(["변동률", "기존가격"])
        else:
            new_row.extend(["", ""])  # Placeholder values
        
        # Process existing data
        for cell in row:
            new_price = cell.value
            
            # Price processing (F column, row 6 and above)
            if cell.column == PRICE_COLUMN + 1 and cell.row >= DATA_START_ROW:  # +1 for 1-indexed
                product_name = row[PRODUCT_NAME_COLUMN].value if row[PRODUCT_NAME_COLUMN].value else None
                
                if product_name is not None:
                    try:
                        original_price = float(cell.value) if cell.value else 0
                        new_price, change_text = PriceProcessor.update_price(str(product_name), original_price)
                        
                        # Update A and B columns
                        new_row[0] = change_text
                        new_row[1] = int(original_price)
                        price_info = (original_price, new_price)
                        
                        # Store change information
                        change_info = {
                            'row': row_idx,
                            'product_name': str(product_name),
                            'original_price': original_price,
                            'new_price': new_price,
                            'change': change_text
                        }
                        
                    except (TypeError, ValueError):
                        new_row[0] = "0원"
                        new_row[1] = int(original_price) if isinstance(cell.value, (int, float)) else 0
            
            new_row.append(new_price)
        
        return new_row, price_info, change_info


def clean_dataframe_for_json(df):
    """Clean DataFrame for JSON serialization"""
    df_clean = df.copy()
    df_clean = df_clean.replace([np.inf, -np.inf], np.nan)
    
    numeric_columns = df_clean.select_dtypes(include=[np.number]).columns
    df_clean[numeric_columns] = df_clean[numeric_columns].fillna(0)
    
    string_columns = df_clean.select_dtypes(include=['object']).columns
    df_clean[string_columns] = df_clean[string_columns].fillna('')
    
    date_columns = df_clean.select_dtypes(include=['datetime64']).columns
    for col in date_columns:
        df_clean[col] = df_clean[col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
    
    return df_clean


# API Views
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def hello_rest_api(request):
    """Simple hello endpoint"""
    data = {'message': 'Hello, REST API!'}
    print(data)
    return Response(data)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_excel(request):
    """Upload and preview Excel file"""
    if 'excel_file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        serializer = ExcelDataSerializer()
        sample_data = [serializer.to_representation(row.to_dict()) for _, row in df.head(10).iterrows()]
        
        return Response({
            'message': 'File uploaded successfully',
            'total_rows': len(df),
            'columns': df.columns.tolist(),
            'sample_data': sample_data,
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def process_excel_and_download(request):
    """Process Excel file and return download response"""
    if 'excel_file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['excel_file']
    
    try:
        new_wb, _ = ExcelProcessor.process_workbook(excel_file)
        
        # Save to memory
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        
        # Create download response
        file_name = f"modified_{os.path.splitext(excel_file.name)[0]}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        response['Content-Length'] = str(len(output.getvalue()))
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def process_excel_with_preview(request):
    """Process Excel file and return preview with download link"""
    if 'excel_file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['excel_file']
    
    try:
        new_wb, price_changes = ExcelProcessor.process_workbook(excel_file)
        
        # Save to memory for preview
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        
        # Generate preview data
        df_preview = pd.read_excel(BytesIO(output.getvalue()))
        df_preview_clean = clean_dataframe_for_json(df_preview)
        
        return Response({
            'message': 'File processed successfully with price search',
            'column_modified': 'Column 6 prices updated based on Naver API search (from row 6 onwards)',
            'price_changes': price_changes,
            'total_rows': len(df_preview),
            'columns': df_preview.columns.tolist(),
            'modified_sample': df_preview_clean.head(10).to_dict('records'),
            'color_legend': {
                'green': '1000원 이하 차이',
                'blue': '2000원 이하 차이', 
                'yellow': '3000원 이하 차이',
                'red': '3000원 초과 차이'
            },
            'download_url': '/api/process-excel-download/',
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)